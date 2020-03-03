# !/usr/bin/env python
# -*- coding:utf-8 -*-
# author:LPE
# datetime:2020-02-27 17:11
# description:用于解析excel的方法

import openpyxl
from openpyxl.styles import Font
import time


class ParseExcel(object):

    def __init__(self):
        self.workbook = None
        self.excelFile = None
        self.font = Font(color=None)  # 设置字体颜色
        # 颜色对应的RGB值
        self.RGBDict = {'red': 'FFFF3030', 'green': 'FF008B00'}

    def loadWorkBook(self, excelPathAndName):
        # 将Excel文件加载到内存，并获取其workbook对象
        if not self.workbook:   # 如果是第一次调用，加载到内存，否则直接返回之前保存的数据
            try:
                self.workbook = openpyxl.load_workbook(excelPathAndName)
            except Exception as e:
                raise e
        self.excelFile = excelPathAndName
        return self.workbook

    def getSheetByName(self, sheetName):
        # 根据 sheet名获取该 sheet对象
        try:
            # sheet = self.workbook.get_sheet_by_name(sheetName)
            sheet = self.workbook[sheetName]
            return sheet
        except Exception as e:
            raise e

    def getSheetByIndex(self, sheetIndex):
        # 根据 sheet的索引号获取该 sheet对象
        try:
            sheetname = self.workbook.get_sheet_names()[sheetIndex]
        except Exception as e:
            raise e
        sheet = self.workbook.get_sheet_by_name(sheetname)
        return sheet

    def getRowsNumber(self, sheet):
        # 获取sheet中有数据区域的结束行号

        return sheet.max_row

    def getColsNumber(self, sheet):
        # 获取sheet中有数据区域的结束列号

        return sheet.max_column

    def getStartRowNumber(self, sheet):
        # 获取sheet中有数据区域的开始的行号
        return sheet.min_row

    def getStartColNumber(self, sheet):
        # 获取sheet中有数据区域的开始的列号
        return sheet.min_column

    def getRow(self, sheet, rowNo):
        # 获取sheet中某一行，返回的是这一行所有的数据内容组成的tuple，

        # 下标从1开始，sheet.row[1]表示第一行

        try:
            return list(sheet.rows)[rowNo - 1]
        except Exception as e:
            raise e

    def getColumn(self, sheet, colNo):
        # 获取sheet中某一列，返回的是这一列所有的数据内容组成的tuple，

        # 下标从1开始，sheet.columns[1]表示第一列

        try:
            return list(sheet.columns)[colNo - 1]
        except Exception as e:
            raise e

    def getCellOfValue(self, sheet, coordinate=None, rowNo=None, colNo=None):
        # 根据单元格所在的位置索引获取该单元格中的值，下标从1开始

        # sheet.cell(row = 1, column =1).value, 表示Excel中第一行第一列的值

        if coordinate is not None:
            try:
                return sheet.cell(coordinate=coordinate).value
            except Exception as e:
                raise e
        elif coordinate is None and rowNo is not None and colNo is not None:
            try:
                return sheet.cell(row=rowNo, column=colNo).value
            except Exception as e:
                raise e
        else:
            raise Exception("Insufficient Coordinates of cell!")

    def getCellOfObject(self, sheet, coordinate=None, rowNo=None, colNo=None):
        # 获取某个单元格的对象，可以根据单元格所在位置的数字索引，

        # 也可以直接根据Excel中单元格的编码及坐标
        # 如getCellObject(sheet, coordinate = 'A1') or
        # getCellObject(sheet, rowNo =1, colNo = 2)
        if coordinate is not None:
            try:
                return sheet.cell(coordinate=coordinate)
            except Exception as e:
                raise e
        elif coordinate is None and rowNo is not None and colNo is not None:
            try:
                return sheet.cell(row=rowNo, column=colNo)
            except Exception as e:
                raise e
        else:
            raise Exception("Insufficient Coordinates of cell!")

    def writeCell(self, sheet, content, coordinate=None, rowNo=None, colNo=None, style=None):
        # 根据单元格在Excel中的编码坐标或者数字索引坐标向单元格中写入数据，

        # 下标从1开始，参数style表示字体的颜色的名字，比如red，green
        if coordinate is not None:
            try:
                sheet.cell(coordinate=coordinate).value = content
                if style is not None:
                    sheet.cell(coordinate=coordinate).font = Font(color=self.RGBDict[style])
                self.workbook.save(self.excelFile)
            except Exception as e:
                raise e
        elif coordinate is None and rowNo is not None and colNo is not None:
            try:
                sheet.cell(row=rowNo, column=colNo).value = content
                if style:
                    sheet.cell(row=rowNo, column=colNo).font = Font(color=self.RGBDict[style])
                self.workbook.save(self.excelFile)
            except Exception as e:
                raise e
        else:
            raise Exception("Insufficient Coordinates of cell!")

    def writeCellCurrentTime(self, sheet, coordinate=None, rowNo=None, colNo=None):
        # 写入当前的时间，下标从1开始

        now = int(time.time())  # 显示为时间戳
        timeArray = time.localtime(now)
        currentTime = time.strftime("%Y-%m-%d %H:%M:%S", timeArray)
        if coordinate is not None:
            try:
                sheet.cell(coordinate=coordinate).value = currentTime
                self.workbook.save(self.excelFile)
            except Exception as e:
                raise e
        elif coordinate is None and rowNo is not None and colNo is not None:
            try:
                sheet.cell(row=rowNo, column=colNo).value = currentTime
                self.workbook.save(self.excelFile)
            except Exception as e:
                raise e
        else:
            raise Exception("Insufficient Coordinates of cell!")


if __name__ == '__main__':
    pe = ParseExcel()
    # 测试所用的Excel文件“testData.xlsx”请自行创建
    # pe.loadWorkBook("C:\\Users\\FHQI\\Documents\\PycharmProjects\\kh_web_test\\testData\\testData.xlsx")
    pe.loadWorkBook("E:\\PyCharmProject\\frame_demo\\testData\\testData.xlsx")
    print("通过名称获取sheet对象的名字：", pe.getSheetByName("账号").title)
    print("通过index序号获取sheet对象的名字：", pe.getSheetByIndex(0).title)
    # sheet = pe.getSheetByIndex(0)
    sheet = pe.getSheetByName("账号")
    print(type(sheet))
    print(pe.getRowsNumber(sheet))  # 获取最大行号

    print(pe.getColsNumber(sheet))  # 获取最大列号

    rows = pe.getRow(sheet, 2)  # 获取第二行

    for i in rows:
        print(i.value)
    # 获取第一行第一列单元格内容
    print(pe.getCellOfValue(sheet, rowNo=1, colNo=1))
    pe.writeCell(sheet, "我在易联众", rowNo=10, colNo=10)
    pe.writeCellCurrentTime(sheet, rowNo=10, colNo=11)